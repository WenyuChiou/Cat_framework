"""Generate CAT411 Team Task Board Excel — clean English version with diagrams."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DOCS_DIR = Path(__file__).parent

# --- Styles ---
H_FONT = Font(name="Consolas", bold=True, size=11, color="FFFFFF")
H_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
BOLD = Font(name="Calibri", bold=True, size=10)
NORM = Font(name="Calibri", size=10)
MONO = Font(name="Consolas", size=9)
MONO_B = Font(name="Consolas", bold=True, size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
BDR = Border(*(Side(style="thin"),) * 4)
NO_BDR = Border()

PRIO = {"P0": "FCE4EC", "P1": "FFF3E0", "P2": "E8F5E9"}
STAT = {"TODO": "F5F5F5", "IN_PROGRESS": "FFF9C4", "DONE": "C8E6C9",
        "BLOCKED": "FFCDD2", "IN_USE": "C8E6C9", "PARTIAL": "FFF9C4"}
WK = {"PLANNED": "E3F2FD", "ACTIVE": "FFF9C4", "DONE": "C8E6C9", "BLOCKED": "FFCDD2"}
LAYER_CLR = {"L0": "EDE7F6", "L1": "E3F2FD", "L2": "E8F5E9",
             "L3": "FFF3E0", "L4": "FCE4EC", "Entry": "F5F5F5"}


def _f(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _header(ws, cols, widths, start_row=1):
    for i, (h, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=start_row, column=i, value=h)
        c.font, c.fill, c.alignment, c.border = H_FONT, H_FILL, CTR, BDR
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{start_row + 1}"


def _row(ws, r, vals, bold=False):
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = BOLD if bold else NORM
        c.alignment = WRAP
        c.border = BDR


def _title_bar(ws, row, col_span, text, fill=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = TITLE_FONT
    c.fill = fill or H_FILL
    c.alignment = CTR


def _diagram_block(ws, start_row, lines, col_span=6):
    """Write a monospace diagram block spanning multiple merged rows."""
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=col_span)
    c = ws.cell(row=start_row, column=1, value="\n".join(lines))
    c.font = MONO
    c.alignment = LEFT_TOP
    c.fill = _f("F8F9FA")
    ws.row_dimensions[start_row].height = max(15 * len(lines), 30)
    return start_row + 1


# =====================================================================
# Sheet 1: Architecture
# =====================================================================
def sheet_architecture(wb):
    ws = wb.active
    ws.title = "Architecture"

    # --- Section A: Pipeline Diagram (image or ASCII fallback) ---
    pipeline_img = DOCS_DIR / "diagram_pipeline.png"
    dependency_img = DOCS_DIR / "diagram_dependency.png"

    next_row = 1
    if pipeline_img.exists():
        _title_bar(ws, 1, 6, "CAT411 FRAMEWORK ARCHITECTURE")
        img = XlImage(str(pipeline_img))
        img.width, img.height = 900, 420
        ws.add_image(img, "A2")
        # Reserve rows for the image
        ws.row_dimensions[2].height = 320
        next_row = 3
    else:
        _title_bar(ws, 1, 6, "CAT411 FRAMEWORK ARCHITECTURE")
        pipeline = [
            "",
            "  BACKBONE PIPELINE (end-to-end CAT model)",
            "  ==========================================",
            "",
            "  +----------------+     +----------------+     +-------------------+     +----------------+     +----------------+",
            "  |    HAZARD       | --> |   EXPOSURE     | --> |  VULNERABILITY    | --> |     LOSS       | --> |  VISUALIZATION |",
            "  |  (Ground Motion)|     | (Bridge Assets)|     | (Fragility Curves)|     | ($ & Downtime) |     |  (Maps & Plots)|",
            "  +----------------+     +----------------+     +-------------------+     +----------------+     +----------------+",
            "         |                      |                        |                        |                      |",
            "    ShakeMap/GMPE         NBI -> HWB class         P(DS|IM) per          E[Loss], EP curve         Dashboard,",
            "    IM at each bridge     + replacement cost       bridge & DS           AAL, loss ratio          risk maps",
            "",
            "  Data flow:  grid.xml / GMPE  --->  Sa(g) per bridge  --->  P(DS)  --->  E[L] = sum(P * DR * Cost)  --->  PNG + CSV",
            "",
        ]
        next_row = _diagram_block(ws, 2, pipeline)

    # --- Section B: Module Dependency Diagram (image or ASCII fallback) ---
    if dependency_img.exists():
        _title_bar(ws, next_row, 6, "MODULE DEPENDENCY (5 LAYERS)", _f("1F4E79"))
        next_row += 1
        img2 = XlImage(str(dependency_img))
        img2.width, img2.height = 900, 500
        ws.add_image(img2, f"A{next_row}")
        ws.row_dimensions[next_row].height = 380
        next_row += 1
    else:
        _title_bar(ws, next_row, 6, "MODULE DEPENDENCY (5 LAYERS)", _f("1F4E79"))
        next_row += 1
        dep_diagram = [
            "",
            "  L0 Data        hazus_params ------------------------------------+",
            "                 config                                           |",
            "                   |                                              |",
            "  L1 Core        hazard          bridge_classes <-- hazus_params  |",
            "                   |  \\              |                             |",
            "                   |   SiteParams    classify_bridge()            |",
            "                   |                 |                             |",
            "  L2 Domain      exposure <------ hazard      data_loader <-- bridge_classes",
            "                   |               fragility <-- hazus_params     |",
            "                   |                 |                             |",
            "  L3 Pipeline    engine <-- hazard + exposure + loss              |",
            "                              loss <-- fragility <----------------+",
            "                                |",
            "  L4 Output      plotting <-- fragility     northridge_case <-- fragility",
            "                 hazard_download",
            "                   |",
            "  Entry           main.py  <--  ALL modules",
            "",
            "  Arrow meaning:  A <-- B  =  'A imports from B'",
            "",
        ]
        next_row = _diagram_block(ws, next_row, dep_diagram)

    # --- Section C: Module Table ---
    _title_bar(ws, next_row, 6, "MODULE REFERENCE TABLE", _f("1F4E79"))
    next_row += 1

    cols = ["Layer", "Module", "Role", "Key API", "Upstream", "Downstream"]
    widths = [8, 18, 32, 38, 22, 22]
    for i, (h, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=next_row, column=i, value=h)
        c.font = H_FONT
        c.fill = _f("1F4E79")
        c.alignment = CTR
        c.border = BDR
        ws.column_dimensions[get_column_letter(i)].width = w
    next_row += 1

    modules = [
        ("L0", "hazus_params", "Fragility parameters (14 HWB classes)", "HAZUS_BRIDGE_FRAGILITY", "—", "bridge_classes, fragility, loss"),
        ("L0", "config", "YAML config loader", "load_config(), AnalysisConfig", "—", "main"),
        ("L1", "hazard", "GMPE + ground motion fields", "boore_atkinson_2008_sa10(), generate_ground_motion_fields()", "—", "engine, exposure"),
        ("L1", "bridge_classes", "NBI -> HWB classification", "classify_bridge(), get_bridge_params()", "hazus_params", "data_loader"),
        ("L1", "interpolation", "Spatial interpolation (5 methods)", "interpolate_im()", "—", "main"),
        ("L2", "fragility", "Fragility curves + skew modification", "damage_state_probabilities(), apply_skew_modification()", "hazus_params", "loss, plotting, northridge_case"),
        ("L2", "exposure", "Portfolio + replacement cost", "create_portfolio_from_nbi(), portfolio_to_sites()", "hazard", "engine"),
        ("L2", "data_loader", "Parse ShakeMap / NBI / stations", "parse_shakemap_grid(), parse_nbi(), load_all()", "bridge_classes", "main"),
        ("L3", "loss", "Loss, EP curve, AAL", "compute_portfolio_loss(), compute_ep_curve(), compute_aal()", "fragility", "engine"),
        ("L3", "engine", "Deterministic + probabilistic runner", "run_deterministic(), run_probabilistic()", "hazard, exposure, loss", "main"),
        ("L4", "plotting", "14+ visualizations", "plot_analysis_summary(), plot_bridge_damage_map()", "fragility", "main"),
        ("L4", "northridge_case", "1994 Northridge validation data", "compare_predicted_vs_observed()", "fragility", "main"),
        ("L4", "hazard_download", "USGS data downloader", "download_all_hazard_data()", "—", "main"),
        ("Entry", "main.py", "CLI entry point (9 analysis modes)", "run_data_analysis(), run_full_analysis()", "ALL", "(user)"),
    ]
    for row_data in modules:
        _row(ws, next_row, row_data)
        lk = row_data[0]
        clr = LAYER_CLR.get(lk, "F5F5F5")
        for j in range(1, len(cols) + 1):
            ws.cell(row=next_row, column=j).fill = _f(clr)
        next_row += 1


# =====================================================================
# Sheet 2: Module_Tasks
# =====================================================================
def sheet_tasks(wb):
    ws = wb.create_sheet("Module_Tasks")
    cols = ["ID", "Module", "Task", "P", "Assignee", "Status", "Blocked By", "Week", "Notes"]
    _header(ws, cols, [7, 14, 44, 5, 10, 12, 12, 6, 32])

    tasks = [
        # ── W1 (2/19-2/25): Hazard & Data ──
        ("T01", "hazard", "Review Hazus params for bridge types (Table 7.9 fragility)", "P0",
         "Kubilay, Michel", "IN_PROGRESS", "", "W1", "Get Sa(1.0s) at 4,853 bridge sites"),
        ("T02", "data_loader", "NBI data quality check (missing values, coords, classification)", "P0",
         "Sirisha", "IN_PROGRESS", "", "W1", "Confirm bridge data is clean"),
        ("T03", "hazus_params", "Spot-check median & beta values against HAZUS manual", "P0",
         "Sirisha", "IN_PROGRESS", "", "W1", "Table 7.9 verification"),
        ("T04", "ALL", "Code + GitHub collaboration setup; leave blanks per module", "P0",
         "Wenyu, Anik", "IN_PROGRESS", "", "W1", "Ongoing"),
        ("T05", "ALL", "Biweekly presentation slides", "P0",
         "Kubilay, Michel", "TODO", "", "W1", "Summarize all work"),
        ("T06", "data_loader", "Verify %g -> g conversion consistency", "P0",
         "Wenyu", "DONE", "", "W1", "Confirmed: parse_shakemap divides by 100"),
        ("T07", "fragility", "Fragility verification (monotonicity / bounds / sum)", "P0",
         "Wenyu", "DONE", "", "W1", "Passed --fragility-only"),
        ("T08", "hazard", "Stabilize im_source decision flow (shakemap first)", "P0",
         "", "TODO", "", "W1", "main.py _compute_bridge_damage"),
        ("T09", "interpolation", "Benchmark interpolation methods (nearest vs IDW)", "P0",
         "", "TODO", "T08", "W1", "Compare on Northridge ShakeMap"),
        ("T10", "ALL", "Integration test: run_data_analysis end-to-end", "P0",
         "Wenyu", "TODO", "T02,T08", "W1", "M1 acceptance gate"),
        # ── W2 (2/26-3/4): Vulnerability & Loss ──
        ("T11", "plotting", "Plot fragility curves for all 14 HWB classes", "P0",
         "", "TODO", "T03", "W2", "Visual check: S-shaped, ordered"),
        ("T12", "bridge_classes", "Bridge classification check (NBI -> HWB mapping logic)", "P0",
         "", "TODO", "T02", "W2", "Correct class assignment"),
        ("T13", "engine", "Run full pipeline (hazard -> fragility -> loss) E2E", "P0",
         "", "TODO", "T10", "W2", "First results for all 4,853 bridges"),
        ("T14", "loss", "Validate expected_loss & loss_ratio reasonableness", "P0",
         "", "TODO", "T13", "W2", "Define acceptable range"),
        ("T15", "loss", "Verify count_by_ds sums to bridge count", "P0",
         "", "TODO", "T14", "W2", ""),
        ("T16", "engine", "Deterministic fixed-seed regression test", "P0",
         "", "TODO", "T14", "W2", "Drift threshold TBD"),
        ("T17", "hazard", "Collect multi-period GMPE coefficients (PGA, SA0.3, SA3.0)", "P1",
         "", "TODO", "", "W2", "Boore & Atkinson 2008 Tables"),
        ("T18", "hazard", "Obtain Vs30 grid for study region", "P1",
         "", "TODO", "", "W2", "USGS / Wills et al."),
        ("T19", "fragility", "E2E test of fragility_overrides from config.yaml", "P1",
         "", "TODO", "", "W2", "Mechanism exists, untested"),
        ("T20", "ALL", "Unit tests: hazard / fragility / loss core functions", "P1",
         "", "TODO", "", "W2", "pytest structure TBD"),
        # ── W3 (3/5-3/11): Validation & Functionality ──
        ("T21", "northridge", "Compare model output vs actual Northridge damage records", "P0",
         "", "TODO", "T13", "W3", "Match 170 real cases?"),
        ("T22", "engine", "Design functionality module (serviceability + recovery)", "P1",
         "", "TODO", "T13", "W3", "Can bridge carry traffic? Repair time?"),
        ("T23", "config", "Calibration factor sensitivity test (ON/OFF comparison)", "P1",
         "", "TODO", "T21", "W3", ""),
        ("T24", "hazard", "Implement GMPE path (currently NotImplementedError)", "P1",
         "", "TODO", "T17,T18", "W3", "Needs multi-period + Vs30"),
        ("T25", "loss", "Update damage ratio table (HAZUS MR4 -> MR5)", "P1",
         "", "TODO", "", "W3", "Compare versions"),
        ("T26", "exposure", "Define skew angle data source + replacement cost update", "P1",
         "", "TODO", "", "W3", ""),
        ("T27", "plotting", "Freeze dashboard indicator set", "P2",
         "", "TODO", "", "W3", "plot_analysis_summary metrics"),
        # ── W4 (3/12-3/18): Report ──
        ("T28", "report", "Write results section (figures, tables, analysis)", "P0",
         "ALL", "TODO", "T21", "W4", "Each person writes their module"),
        ("T29", "report", "Integrate all sections + final review", "P0",
         "ALL", "TODO", "T28", "W4", "Complete draft of full report"),
        ("T30", "plotting", "Report templates (technical + management)", "P2",
         "", "TODO", "T27", "W4", ""),
    ]
    for i, row in enumerate(tasks, 2):
        _row(ws, i, row)
        p = row[3]
        if p in PRIO:
            ws.cell(row=i, column=4).fill = _f(PRIO[p])
        s = row[5]
        if s in STAT:
            ws.cell(row=i, column=6).fill = _f(STAT[s])


# =====================================================================
# Sheet 3: Data_Flow
# =====================================================================
def sheet_data_flow(wb):
    ws = wb.create_sheet("Data_Flow")

    # --- Diagram at top (image or ASCII fallback) ---
    dataflow_img = DOCS_DIR / "diagram_dataflow.png"
    if dataflow_img.exists():
        _title_bar(ws, 1, 8, "DATA FLOW OVERVIEW")
        img = XlImage(str(dataflow_img))
        img.width, img.height = 900, 420
        ws.add_image(img, "A2")
        ws.row_dimensions[2].height = 320
        next_row = 3
    else:
        _title_bar(ws, 1, 8, "DATA FLOW OVERVIEW")
        flow_lines = [
            "",
            "  +-----------+    grid.xml     +-----------+   Sa(g) array   +-----------+   {DS: prob}   +-----------+   E[L], AAL   +-----------+",
            "  |  DATA     | -------------> |  HAZARD   | -------------> | FRAGILITY | -------------> |   LOSS    | ------------> |  OUTPUT   |",
            "  |  LOADER   |    NBI .txt     |  MODULE   |   per bridge    |  MODULE   |   per bridge   |  MODULE   |   EP curve    |  MODULE   |",
            "  +-----------+ -------------> +-----------+                 +-----------+                +-----------+               +-----------+",
            "       |                             |                             |                            |                          |",
            "   parse_shakemap_grid()      interpolate_im()            damage_state_probs()        compute_portfolio_loss()     plot_*(), CSV",
            "   parse_nbi()                boore_atkinson_2008()       apply_skew_mod()            compute_ep_curve()           dashboard",
            "   classify_nbi_to_hazus()                                                            compute_aal()",
            "",
            "  KEY DATA OBJECTS:",
            "  -----------------",
            "  ShakeMap grid  :  DataFrame [lat, lon, pga, sa03, sa10, sa30]    (units: g)",
            "  NBI bridges    :  DataFrame [structure_number, lat, lon, hwb_class, material, deck_area, ...]",
            "  BridgeExposure :  dataclass [bridge_id, lat, lon, hwb_class, replacement_cost, vs30, skew_angle]",
            "  SiteParams     :  dataclass [lat, lon, vs30]",
            "  BridgeLossResult    :  [bridge_id, sa, damage_probs, expected_loss, loss_ratio]",
            "  PortfolioLossResult :  [total_loss, loss_ratio, loss_by_class, count_by_ds, AAL]",
            "",
        ]
        next_row = _diagram_block(ws, 2, flow_lines, 8)

    # --- Matrix below ---
    _title_bar(ws, next_row, 8, "MODULE-TO-MODULE DATA MATRIX",
               _f("1F4E79"))
    next_row += 1

    mods = ["hazus_params", "bridge_classes", "hazard", "exposure",
            "data_loader", "fragility", "interpolation", "loss",
            "engine", "plotting", "main"]

    c = ws.cell(row=next_row, column=1, value="FROM \\ TO")
    c.font, c.fill, c.alignment, c.border = H_FONT, H_FILL, CTR, BDR
    ws.column_dimensions["A"].width = 16
    for j, m in enumerate(mods, 2):
        c = ws.cell(row=next_row, column=j, value=m)
        c.font, c.fill, c.alignment, c.border = H_FONT, H_FILL, CTR, BDR
        ws.column_dimensions[get_column_letter(j)].width = 26
    next_row += 1

    for i, m in enumerate(mods):
        c = ws.cell(row=next_row + i, column=1, value=m)
        c.font, c.fill, c.border = BOLD, _f("D6E4F0"), BDR

    flows = {
        ("hazus_params", "bridge_classes"): "HAZUS_BRIDGE_FRAGILITY dict",
        ("hazus_params", "fragility"): "HAZUS_BRIDGE_FRAGILITY,\nDAMAGE_STATE_ORDER",
        ("hazus_params", "loss"): "DAMAGE_STATE_ORDER",
        ("hazus_params", "plotting"): "HAZUS_BRIDGE_FRAGILITY",
        ("bridge_classes", "data_loader"): "classify_bridge()\n-> HWB class string",
        ("hazard", "engine"): "EarthquakeScenario,\nSa(g) arrays (n_sites x n_real)",
        ("hazard", "exposure"): "SiteParams type import",
        ("exposure", "engine"): "list[BridgeExposure],\nlist[SiteParams]",
        ("data_loader", "main"): "DataFrames:\nshakemap grid, NBI, stations",
        ("fragility", "loss"): "damage_state_probabilities()\n-> {DS: float}",
        ("fragility", "plotting"): "compute_all_curves()\n-> {DS: np.ndarray}",
        ("interpolation", "main"): "np.ndarray:\nIM values at bridge sites",
        ("loss", "engine"): "PortfolioLossResult,\nEP curve dict, AAL float",
        ("engine", "main"): "DeterministicResult /\nProbabilisticResult",
        ("plotting", "main"): "PNG file paths (str)",
    }
    idx = {m: i for i, m in enumerate(mods)}
    for (s, t), desc in flows.items():
        r = next_row + idx[s]
        cc = idx[t] + 2
        cell = ws.cell(row=r, column=cc, value=desc)
        cell.font, cell.alignment, cell.border = NORM, WRAP, BDR
        cell.fill = _f("E8F5E9")

    for i in range(len(mods)):
        for j in range(len(mods)):
            cell = ws.cell(row=next_row + i, column=j + 2)
            cell.border = BDR
            if cell.value is None and i == j:
                cell.value, cell.alignment, cell.fill = "—", CTR, _f("BDBDBD")

    ws.freeze_panes = f"B{next_row}"


# =====================================================================
# Sheet 4: Literature
# =====================================================================
def sheet_literature(wb):
    ws = wb.create_sheet("Literature")
    cols = ["ID", "Topic", "Reference", "Status", "Assigned To", "Module"]
    _header(ws, cols, [7, 18, 48, 12, 12, 18])

    items = [
        ("L01", "GMPE", "Boore & Atkinson (2008) — multi-period coefficients", "PARTIAL", "", "hazard"),
        ("L02", "GMPE", "Campbell & Bozorgnia (2008) — alternative NGA GMPE", "TODO", "", "hazard"),
        ("L03", "Vs30", "Wills et al. (2015) / USGS Vs30 grid", "TODO", "", "hazard"),
        ("L04", "Fragility", "HAZUS-MH MR4/MR5 Technical Manual Ch.7", "IN_USE", "", "hazus_params"),
        ("L05", "Fragility", "Shinozuka et al. (2000) — empirical bridge fragility", "TODO", "", "fragility"),
        ("L06", "Damage Ratio", "HAZUS Table 7.9 + Werner (2006) REDARS", "PARTIAL", "", "loss"),
        ("L07", "Cost", "RSMeans Heavy Construction Cost Data 2024", "TODO", "", "exposure"),
        ("L08", "Correlation", "Jayaram & Baker (2009) — spatial correlation model", "DONE", "", "hazard"),
        ("L09", "Validation", "Basoz & Kiremidjian (1998) — Northridge bridge damage", "PARTIAL", "", "northridge_case"),
        ("L10", "Classification", "NBI Coding Guide (FHWA)", "IN_USE", "", "bridge_classes"),
        ("L11", "Skew", "HAZUS Technical Manual S7.2.4 — skew factor", "PARTIAL", "", "fragility"),
        ("L12", "Hazard Map", "USGS NSHM 2023 — probabilistic hazard curves", "TODO", "", "hazard_download"),
        ("L13", "CAT Theory", "Grossi & Kunreuther (2005) Catastrophe Modeling", "TODO", "", "engine"),
        ("L14", "Fragility-TW", "NCREE Bridge Fragility Reports (Taiwan)", "TODO", "", "fragility"),
    ]
    for i, row in enumerate(items, 2):
        _row(ws, i, row)
        s = row[3]
        if s in STAT:
            ws.cell(row=i, column=4).fill = _f(STAT[s])


# =====================================================================
# Sheet 5: Timeline
# =====================================================================
def sheet_timeline(wb):
    ws = wb.create_sheet("Timeline")

    # --- Milestone diagram ---
    _title_bar(ws, 1, 7, "MILESTONES & TIMELINE")
    ms_lines = [
        "",
        "  TEAM: Wenyu (lead/code) | Kubilay & Michel (hazard/params) | Sirisha (data QA) | Anik (code)",
        "",
        "  Phase 1: HAZARD         Phase 2: VULNERABILITY      Phase 3: VALIDATION       Phase 4: REPORT",
        "  (data + params)         (fragility + loss E2E)      (Northridge + calibrate)   (write + integrate)",
        "  ------------------      ----------------------      -----------------------    -------------------",
        "  Review Hazus params     Plot 14 HWB curves          Model vs observed damage   Results section",
        "  NBI data quality        HWB mapping check           Functionality module       Integrate + review",
        "  Code/Git setup          Run full pipeline E2E       Sensitivity tests          Final draft",
        "",
        "  |--- W1: 2/19-2/25 ---|--- W2: 2/26-3/4 ---|--- W3: 3/5-3/11 ---|--- W4: 3/12-3/18 ---|",
        "  |  Phase 1: Hazard    |  Phase 2: Vuln.    |  Phase 3: Valid.   |  Phase 4: Report    |",
        "  |  << WE ARE HERE >>  |                    |                    |                     |",
        "",
    ]
    next_row = _diagram_block(ws, 2, ms_lines, 7)

    # --- Gantt-like table ---
    cols = ["ID", "Task", "P",
            "W1\n2/19-2/25", "W2\n2/26-3/4", "W3\n3/5-3/11", "W4\n3/12-3/18"]
    for i, (h, w) in enumerate(zip(cols, [7, 40, 5, 14, 14, 14, 14]), 1):
        c = ws.cell(row=next_row, column=i, value=h)
        c.font, c.fill, c.alignment, c.border = H_FONT, H_FILL, CTR, BDR
        ws.column_dimensions[get_column_letter(i)].width = w
    next_row += 1

    tl = [
        # W1 tasks
        ("T01", "Review Hazus params (Kubilay, Michel)", "P0", "ACTIVE", "", "", ""),
        ("T02", "NBI data quality check (Sirisha)", "P0", "ACTIVE", "", "", ""),
        ("T03", "Spot-check median & beta (Sirisha)", "P0", "ACTIVE", "", "", ""),
        ("T04", "Code + GitHub collab (Wenyu, Anik)", "P0", "ACTIVE", "ACTIVE", "ACTIVE", "ACTIVE"),
        ("T05", "Biweekly presentation (Kubilay, Michel)", "P0", "ACTIVE", "", "", ""),
        ("T06", "%g->g conversion (Wenyu)", "P0", "DONE", "", "", ""),
        ("T07", "Fragility verification (Wenyu)", "P0", "DONE", "", "", ""),
        ("T08", "im_source stabilization", "P0", "ACTIVE", "", "", ""),
        ("T10", "Integration test E2E (Wenyu)", "P0", "ACTIVE", "DONE", "", ""),
        # W2 tasks
        ("T11", "Plot fragility curves (14 HWB)", "P0", "", "ACTIVE", "", ""),
        ("T12", "HWB classification check", "P0", "", "ACTIVE", "", ""),
        ("T13", "Run full pipeline E2E (4,853 bridges)", "P0", "", "ACTIVE", "", ""),
        ("T14", "Loss reasonableness check", "P0", "", "ACTIVE", "", ""),
        ("T17", "Collect GMPE coefficients", "P1", "", "ACTIVE", "", ""),
        ("T20", "Unit tests", "P1", "", "ACTIVE", "ACTIVE", ""),
        # W3 tasks
        ("T21", "Northridge: predicted vs observed", "P0", "", "", "ACTIVE", ""),
        ("T22", "Functionality module (serviceability)", "P1", "", "", "ACTIVE", ""),
        ("T23", "Calibration sensitivity test", "P1", "", "", "ACTIVE", ""),
        ("T24", "GMPE path implementation", "P1", "", "", "ACTIVE", ""),
        # W4 tasks
        ("T28", "Write results section (ALL)", "P0", "", "", "", "ACTIVE"),
        ("T29", "Integrate sections + final review (ALL)", "P0", "", "", "", "ACTIVE"),
    ]
    for row_data in tl:
        _row(ws, next_row, row_data[:3])
        p = row_data[2]
        if p in PRIO:
            ws.cell(row=next_row, column=3).fill = _f(PRIO[p])
        for j, val in enumerate(row_data[3:], 4):
            c = ws.cell(row=next_row, column=j, value=val)
            c.font, c.alignment, c.border = NORM, CTR, BDR
            if val in WK:
                c.fill = _f(WK[val])
        next_row += 1

    # Legend
    next_row += 1
    ws.cell(row=next_row, column=1, value="Legend:").font = BOLD
    for j, (lab, clr) in enumerate([("PLANNED", "E3F2FD"), ("ACTIVE", "FFF9C4"),
                                     ("DONE", "C8E6C9"), ("BLOCKED", "FFCDD2")]):
        c = ws.cell(row=next_row, column=2 + j, value=lab)
        c.fill, c.font, c.alignment, c.border = _f(clr), NORM, CTR, BDR


# =====================================================================
# Sheet 6: Glossary
# =====================================================================
def sheet_glossary(wb):
    ws = wb.create_sheet("Glossary")
    cols = ["Term", "Full Name", "Definition"]
    _header(ws, cols, [14, 36, 60])

    terms = [
        ("CAT", "Catastrophe Model", "Computational framework to estimate losses from natural disasters"),
        ("GMPE", "Ground Motion Prediction Equation", "Empirical equation: IM = f(Mw, distance, Vs30, fault type)"),
        ("IM", "Intensity Measure", "Quantitative ground shaking metric (PGA, SA at various periods)"),
        ("PGA", "Peak Ground Acceleration", "Maximum ground surface acceleration, in units of g"),
        ("SA10", "Spectral Acceleration at 1.0s", "Default IM for HAZUS bridge fragility; primary IM in this framework"),
        ("Vs30", "Shear-wave velocity (top 30m)", "Site amplification proxy, units m/s (e.g., 760 = rock, 360 = stiff soil)"),
        ("NBI", "National Bridge Inventory", "FHWA database of all US bridges with structural attributes"),
        ("HWB", "HAZUS Bridge Class", "Classification HWB1-HWB28 based on material, span type, design era"),
        ("DS", "Damage State", "Discrete levels: none / slight / moderate / extensive / complete"),
        ("Fragility", "Fragility Curve", "P(DS >= ds | IM) modeled as lognormal CDF with median & beta"),
        ("EP", "Exceedance Probability", "Annual probability that loss exceeds a given $ threshold"),
        ("AAL", "Average Annual Loss", "Expected annual loss = integral under EP curve"),
        ("E[L]", "Expected Loss", "Probability-weighted loss across all damage states"),
        ("LR", "Loss Ratio", "E[L] / replacement cost, dimensionless [0, 1]"),
        ("Mw", "Moment Magnitude", "Standard earthquake magnitude scale used in GMPE"),
        ("R_JB", "Joyner-Boore Distance", "Shortest distance from site to fault surface projection (km)"),
        ("IDW", "Inverse Distance Weighting", "Spatial interpolation method, weight proportional to 1/d^p"),
    ]
    for i, row in enumerate(terms, 2):
        _row(ws, i, row)
        if i % 2 == 0:
            for j in range(1, 4):
                ws.cell(row=i, column=j).fill = _f("F5F5F5")


# =====================================================================
# Main
# =====================================================================
def main():
    wb = Workbook()
    sheet_architecture(wb)
    sheet_tasks(wb)
    sheet_data_flow(wb)
    sheet_literature(wb)
    sheet_timeline(wb)
    sheet_glossary(wb)

    out = "docs/CAT411_Team_Task_Board.xlsx"
    wb.save(out)
    print(f"Saved: {out}  |  Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
